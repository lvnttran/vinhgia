from typing import List

from fastapi import APIRouter, Depends, status, Request, File, UploadFile, Form
from fastapi.responses import HTMLResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session

from api import factory
from database import configuration
from schema import schemas
from models.models import *
from schema.oa2 import get_current_user

from openpyxl import load_workbook
import subprocess
import pdfkit
from io import BytesIO
from datetime import datetime

router = APIRouter(tags=["Factory"], prefix="/api/factory")
get_db = configuration.get_db

# Jinja2 Template directory
templates = Jinja2Templates(directory="templates")


@router.get("/", status_code=status.HTTP_200_OK, response_model=List[schemas.FactoryShow])
def get_factory(db: Session = Depends(get_db)):
    return factory.get_all(db)


@router.get("/factorys", status_code=status.HTTP_200_OK, response_model=List[schemas.FactoryModel])
def get_factorys(db: Session = Depends(get_db)):
    return factory.get_all_full(db)


@router.post("/", status_code=status.HTTP_201_CREATED, response_model=schemas.FactoryShow)
def create_factory(request: schemas.Factory, db: Session = Depends(get_db)):
    response = factory.create(request, db)
    return response


@router.get("/{id}", status_code=status.HTTP_200_OK, response_model=schemas.FactoryShow)
def get_factory(id: int, db: Session = Depends(get_db)):
    return factory.show(id, db)


@router.delete("/{id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_factory(
        id: int,
        db: Session = Depends(get_db),
):
    return factory.destroy(id, db)


# DON HANG XEM
@router.get("/donhang/xem", response_class=HTMLResponse, tags=["ĐƠN HÀNG"])
def read_all_donhang(request: Request, db: Session = Depends(get_db)):
    # Retrieve the 50 most recent records from the database
    donhang_list = db.query(Donhang).order_by(Donhang.donhang_id.desc()).limit(50).all()
    return templates.TemplateResponse("donhang_xem.html", {"request": request, "donhang_list": donhang_list})


# DON HANG IMPORT
@router.post("/donhang/xuly", response_class=HTMLResponse, tags=["ĐƠN HÀNG"])
async def process_donhang(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    # Read the content of the file into memory
    content = await file.read()

    # Create a seekable file-like object from the content
    file_obj = BytesIO(content)

    # Load the Excel file
    workbook = load_workbook(file_obj, data_only=True)

    # Get the first sheet
    sheet = workbook.active

    # Initialize an empty list to store the extracted data
    data = []

    # Iterate over the rows starting from the 4th row (row_index=3)
    for row_index, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        # Extract the values from the row
        donhang_madh = row[3]  # Assuming the "MÃ ĐƠN" column is in the 4th position (column D)
        donhang_masp = row[4]  # Assuming the "MÃ SP" column is in the 5th position (column E)
        donhang_mota = row[7]  # Assuming the "MÔ TẢ" column is in the 8th position (column H)
        donhang_soluong = row[6]  # Assuming the "SỐ LƯỢNG" column is in the 7th position (column G)
        donhang_ngay = row[1]  # Assuming the "NGÀY" column is in the 2nd position (column B)

        # Convert the "ngày" column value to a datetime.date object
        ngay_str = row[1]
        if isinstance(ngay_str, datetime):
            donhang_ngay = ngay_str.date()
        else:
            donhang_ngay = datetime.strptime(ngay_str, "%d/%m/%Y").date() if ngay_str else None

        # Create a dictionary to store the extracted values
        donhang_data = {
            'donhang_madh': donhang_madh,
            'donhang_masp': donhang_masp,
            'donhang_mota': donhang_mota,
            'donhang_soluong': donhang_soluong,
            'donhang_ngay': donhang_ngay
        }

        # Append the dictionary to the data list
        data.append(donhang_data)

    # Iterate over the extracted data and save it to the database
    for donhang_data in data:
        donhang = Donhang(**donhang_data)
        db.add(donhang)

    db.commit()

    donhang_list = db.query(Donhang).order_by(Donhang.donhang_id.desc()).limit(50).all()
    return templates.TemplateResponse("donhang_xem.html", {"request": request, "donhang_list": donhang_list})


# MA LOI XEM
@router.get("/maloi/xem", response_class=HTMLResponse, tags=["MÃ LỖI"])
def read_all_donhang(request: Request, db: Session = Depends(get_db)):
    # Retrieve the 50 most recent records from the database
    maloi_list = db.query(Maloi).order_by(Maloi.maloi_id.desc()).limit(50).all()
    return templates.TemplateResponse("maloi_xem.html", {"request": request, "maloi_list": maloi_list})


# MA LOI IMPORT
@router.post("/maloi/xuly", response_class=HTMLResponse, tags=["MÃ LỖI"])
async def process_maloi(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    # Read the content of the file into memory
    content = await file.read()

    # Create a seekable file-like object from the content
    file_obj = BytesIO(content)

    # Load the Excel file
    workbook = load_workbook(file_obj, data_only=True)

    # Get the first sheet
    sheet = workbook.active

    # Initialize an empty list to store the extracted data
    data = []

    # Iterate over the rows starting from the 4th row (row_index=3)
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Extract the values from the row
        maloi_maloi = row[0]  # Assuming the "MÃ LỖI" column is in the 1st position (column A)
        maloi_tenloi = row[1]  # Assuming the "TÊN LỖI" column is in the 2nd position (column B)
        maloi_khacphuc = row[2]  # Assuming the "HƯỚNG KHẮC PHỤC" column is in the 3rd position (column C)
        maloi_nguyennhan = row[3]  # Assuming the "NGUYÊN NHÂN" column is in the 4th position (column D)
        maloi_phongngua = row[4]  # Assuming the "HƯỚNG PHÒNG NGỪA" column is in the 5th position (column E)

        # Create a dictionary to store the extracted values
        maloi_data = {
            'maloi_maloi': maloi_maloi,
            'maloi_tenloi': maloi_tenloi,
            'maloi_khacphuc': maloi_khacphuc,
            'maloi_nguyennhan': maloi_nguyennhan,
            'maloi_phongngua': maloi_phongngua
        }

        # Append the dictionary to the data list
        data.append(maloi_data)

    # Iterate over the extracted data and save it to the database
    for maloi_data in data:
        maloi = Maloi(**maloi_data)
        db.add(maloi)

    db.commit()

    maloi_list = db.query(Maloi).order_by(Maloi.maloi_id.desc()).limit(50).all()
    return templates.TemplateResponse("maloi_xem.html", {"request": request, "maloi_list": maloi_list})


# MA LOI THEM
@router.get("/maloi/themui", response_class=HTMLResponse, tags=["MÃ LỖI"])
async def maloi_them_ui(request: Request):
    return templates.TemplateResponse("maloi_them.html", {"request": request})


@router.post("/maloi/them", response_class=HTMLResponse, tags=["MÃ LỖI"])
def maloi_them(request: Request, maloi_maloi: str = Form(...), maloi_tenloi: str = Form(...),
               maloi_khacphuc: str = Form(...),
               maloi_nguyennhan: str = Form(...), maloi_phongngua: str = Form(...), db: Session = Depends(get_db)):
    maloi = Maloi(maloi_maloi=maloi_maloi, maloi_tenloi=maloi_tenloi, maloi_khacphuc=maloi_khacphuc,
                  maloi_nguyennhan=maloi_nguyennhan, maloi_phongngua=maloi_phongngua)
    db.add(maloi)
    db.commit()
    time.sleep(1)
    maloi_list = db.query(Maloi).order_by(Maloi.maloi_id.desc()).limit(50).all()
    return templates.TemplateResponse("maloi_xem.html", {"request": request, "maloi_list": maloi_list})


# BAOCAOSUCO THEM
@router.get("/baocaosuco/themui", response_class=HTMLResponse, tags=["BÁO CÁO SỰ CỐ"])
async def baocaosuco_them_ui(request: Request, db: Session = Depends(get_db)):
    donhang_list = db.query(Donhang).all()
    maloi_list = db.query(Maloi).all()
    return templates.TemplateResponse("baocaosuco_them.html", {"request": request,
                                                               "donhang_list": donhang_list, "maloi_list": maloi_list})


# BAOCAOSUCO XEM
@router.post("/baocaosuco/them", tags=["BÁO CÁO SỰ CỐ"])
async def baocaosuco_them(request: Request):
    # Get the form data
    form_data = await request.form()

    # Extract the form input values
    madh = form_data["madh"]
    masp = form_data["masp"]
    maloi = form_data["maloi"]
    tenloi = form_data["tenloi"]
    soluongloi = form_data["soluongloi"]
    khacphuc = form_data["khacphuc"]
    nguyennhan = form_data["nguyennhan"]
    phongngua = form_data["phongngua"]
    ccemail = form_data["ccemail"]

    # Generate the PDF content
    pdf_content = f"""
    <h1>Báo Cáo Sự Cố</h1>
    <p>Mã Đơn Hàng: {madh}</p>
    <p>Mã Sản Phẩm: {masp}</p>
    <p>Mã Lỗi: {maloi}</p>
    <p>Tên Lỗi: {tenloi}</p>
    <p>Số Lượng Lỗi: {soluongloi}</p>
    <p>Hướng Khắc Phục: {khacphuc}</p>
    <p>Nguyên Nhân: {nguyennhan}</p>
    <p>Hướng Phòng Ngừa: {phongngua}</p>
    <p>CC Email: {ccemail}</p>
    """

    # Generate the PDF file name
    today = date.today().strftime("%Y_%m_%d")
    file_name = f"{madh}_{masp}_{maloi}.pdf"
    file_path = f"/storage/baocaosuco/{today}/{file_name}"

    # Generate the PDF and save it
    pdfkit.from_string(pdf_content, file_path, options={"quiet": ""})

    return {"message": "PDF generated and saved successfully"}


# BAOCAOSUCO XEM
@router.get("/baocaosuco/xem", tags=["BÁO CÁO SỰ CỐ"])
async def baocaosuco_xem(request: Request):
    # Specify your local directory path
    directory_path = r"E:\1. Research\proposal_smart_fac\om_demo\storage\baocaosuco"

    # Open the file explorer or file manager application with the specified directory path
    subprocess.Popen(f'explorer "{directory_path}"')

    # Return a response to indicate that the operation was successful
    return {"message": "Tổng hợp báo cáo sự cố"}
