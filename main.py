# imports
from pydantic import BaseModel
from fastapi import FastAPI, Request, Response, status, HTTPException, Depends, Form, Cookie, File, UploadFile

from sqlalchemy import create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy import Column, Integer, String, Float, DateTime, ForeignKey, Date
from sqlalchemy.orm import Session, sessionmaker, relationship

from fastapi.security import HTTPBasicCredentials, HTTPBasic
from fastapi.responses import HTMLResponse, RedirectResponse, Response, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from openpyxl import load_workbook
import pandas as pd
from tempfile import TemporaryFile

import pdfkit
import os
import time
from datetime import datetime, timedelta

from passlib.context import CryptContext
from typing import List
from functools import wraps
from enum import Enum
import jwt

from io import BytesIO, BufferedReader
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

# Initialize
app = FastAPI()

# Postgres Database
SQLALCHEMY_DATABASE_URL = "postgresql://postgres:tum19@localhost/factory"
engine = create_engine(SQLALCHEMY_DATABASE_URL)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()


# Model
class Donhang(Base):
    __tablename__ = "donhang"

    donhang_id = Column(Integer, primary_key=True, index=True)
    donhang_madh = Column(String)
    donhang_masp = Column(String)
    donhang_mota = Column(String)
    donhang_soluong = Column(String)
    donhang_ngay = Column(Date)


class Maloi(Base):
    __tablename__ = "maloi"

    maloi_id = Column(Integer, primary_key=True, index=True, autoincrement=True)
    maloi_maloi = Column(String)
    maloi_tenloi = Column(String)
    maloi_khacphuc = Column(String)
    maloi_nguyennhan = Column(String)
    maloi_phongngua = Column(String)


class Fault(Base):
    __tablename__ = "fault"

    fault_id = Column(Integer, primary_key=True, index=True)
    fault_name = Column(String)
    fault_image = Column(String)
    fault_description = Column(String)
    fault_approval = Column(String)

    # Define a one-to-many relationship between Fault and Solution
    solutions = relationship("Solution", back_populates="fault")


class Solution(Base):
    __tablename__ = "solution"

    solution_id = Column(Integer, primary_key=True, index=True)
    solution_detail = Column(String)

    # Define a many-to-one relationship between Solution and Fault
    fault_id = Column(Integer, ForeignKey("fault.fault_id"))
    fault = relationship("Fault", back_populates="solutions")


class User(Base):
    __tablename__ = "user"

    user_id = Column(Integer, primary_key=True, index=True)
    user_name = Column(String)
    user_pass = Column(String)
    user_role = Column(String)


class UserRole(str, Enum):
    admin = "admin"
    client = "client"


# Create tables
Base.metadata.create_all(bind=engine)

# Static file serv
app.mount("/static", StaticFiles(directory="static"), name="static")
# Jinja2 Template directory
templates = Jinja2Templates(directory="templates")

# define the JWT secret key and algorithm
JWT_SECRET_KEY = "mysecretkey"
JWT_ALGORITHM = "HS256"

# create a security object for handling authentication
security = HTTPBasic()
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")


# Dependency
def get_db():
    try:
        db = Session(bind=engine)
        yield db
    finally:
        db.close()


def get_current_user(request: Request, db: Session = Depends(get_db)):
    credentials = HTTPBasicCredentials()
    if request.headers.get("Authorization"):
        scheme, credentials = request.headers["Authorization"].split(" ")
        if scheme.lower() != "basic":
            raise HTTPException(status_code=400, detail="Invalid authentication scheme")
        decoded_credentials = base64.b64decode(credentials).decode("ascii")
        username, password = decoded_credentials.split(":")
        user = authenticate_user(db, username, password)
        if not user:
            raise HTTPException(status_code=401, detail="Invalid username or password")
        return user
    else:
        raise HTTPException(status_code=401, detail="Invalid authorization header")


def check_user_role(role: UserRole):
    def decorator(func):
        @wraps(func)
        def wrapper(*args, **kwargs):
            user = kwargs["user"]
            if user.user_role != role:
                raise HTTPException(status_code=403, detail="You do not have permission to access this resource")
            return func(*args, **kwargs)

        return wrapper

    return decorator


# define the authentication function
def authenticate_user(db: Session, username: str, password: str):
    user = db.query(User).filter(User.user_name == username).first()
    if user and pwd_context.verify(password, user.user_hashed_password):
        return user
    else:
        return None


# define a function to create a JWT access token
def create_access_token(username: str, role: str):
    # calculate the expiration time of the token
    delta = timedelta(hours=1)
    expiration = datetime.utcnow() + delta

    # create the token payload
    payload = {
        "sub": username,
        "role": role,
        "exp": expiration
    }

    # encode the token using the secret key and algorithm
    token = jwt.encode(payload, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)

    return token


# HOME
@app.get("/", response_class=HTMLResponse, tags=["HOME PAGE"])
def home_page(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


# LOGIN
@app.get("/login", response_class=HTMLResponse, tags=["SIGN IN"])
def home_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})


@app.post("/login", response_class=HTMLResponse, tags=["SIGN IN"])
def login(db: Session = Depends(get_db), credentials: HTTPBasicCredentials = Depends(security)):
    user = authenticate_user(db, credentials.username, credentials.password)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    access_token = create_access_token(user.user_name, user.user_role)
    return {"access_token": access_token, "token_type": "bearer"}


# DON HANG XEM
@app.get("/donhang/xem", response_class=HTMLResponse, tags=["ĐƠN HÀNG"])
def read_all_donhang(request: Request, db: Session = Depends(get_db)):
    # Retrieve the 50 most recent records from the database
    donhang_list = db.query(Donhang).order_by(Donhang.donhang_id.desc()).limit(50).all()
    return templates.TemplateResponse("donhang_xem.html", {"request": request, "donhang_list": donhang_list})


# DON HANG IMPORT
@app.post("/donhang/xuly", response_class=HTMLResponse, tags=["ĐƠN HÀNG"])
async def process_donhang(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    # Read the content of the file into memory
    content = await file.read()

    # Create a seekable file-like object from the content
    file_obj = BytesIO(content)

    # Load the Excel file
    workbook = load_workbook(file_obj, data_only=True)

    # Get the first sheet
    sheet = workbook.active

    # Initialize an empty list to store the extracted data
    data = []

    # Iterate over the rows starting from the 4th row (row_index=3)
    for row_index, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        # Extract the values from the row
        donhang_madh = row[3]  # Assuming the "MÃ ĐƠN" column is in the 4th position (column D)
        donhang_masp = row[4]  # Assuming the "MÃ SP" column is in the 5th position (column E)
        donhang_mota = row[7]  # Assuming the "MÔ TẢ" column is in the 8th position (column H)
        donhang_soluong = row[6]  # Assuming the "SỐ LƯỢNG" column is in the 7th position (column G)
        donhang_ngay = row[1]  # Assuming the "NGÀY" column is in the 2nd position (column B)

        # Convert the "ngày" column value to a datetime.date object
        ngay_str = row[1]
        if isinstance(ngay_str, datetime):
            donhang_ngay = ngay_str.date()
        else:
            donhang_ngay = datetime.strptime(ngay_str, "%d/%m/%Y").date() if ngay_str else None

        # Create a dictionary to store the extracted values
        donhang_data = {
            'donhang_madh': donhang_madh,
            'donhang_masp': donhang_masp,
            'donhang_mota': donhang_mota,
            'donhang_soluong': donhang_soluong,
            'donhang_ngay': donhang_ngay
        }

        # Append the dictionary to the data list
        data.append(donhang_data)

    # Iterate over the extracted data and save it to the database
    for donhang_data in data:
        donhang = Donhang(**donhang_data)
        db.add(donhang)

    db.commit()

    donhang_list = db.query(Donhang).order_by(Donhang.donhang_id.desc()).limit(50).all()
    return templates.TemplateResponse("donhang_xem.html", {"request": request, "donhang_list": donhang_list})


# MA LOI XEM
@app.get("/maloi/xem", response_class=HTMLResponse, tags=["MÃ LỖI"])
def read_all_donhang(request: Request, db: Session = Depends(get_db)):
    # Retrieve the 50 most recent records from the database
    maloi_list = db.query(Maloi).order_by(Maloi.maloi_id.desc()).limit(50).all()
    return templates.TemplateResponse("maloi_xem.html", {"request": request, "maloi_list": maloi_list})


# MA LOI IMPORT
@app.post("/maloi/xuly", response_class=HTMLResponse, tags=["MÃ LỖI"])
async def process_maloi(request: Request, file: UploadFile = File(...), db: Session = Depends(get_db)):
    # Read the content of the file into memory
    content = await file.read()

    # Create a seekable file-like object from the content
    file_obj = BytesIO(content)

    # Load the Excel file
    workbook = load_workbook(file_obj, data_only=True)

    # Get the first sheet
    sheet = workbook.active

    # Initialize an empty list to store the extracted data
    data = []

    # Iterate over the rows starting from the 4th row (row_index=3)
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Extract the values from the row
        maloi_maloi = row[0]  # Assuming the "MÃ LỖI" column is in the 1st position (column A)
        maloi_tenloi = row[1]  # Assuming the "TÊN LỖI" column is in the 2nd position (column B)
        maloi_khacphuc = row[2]  # Assuming the "HƯỚNG KHẮC PHỤC" column is in the 3rd position (column C)
        maloi_nguyennhan = row[3]  # Assuming the "NGUYÊN NHÂN" column is in the 4th position (column D)
        maloi_phongngua = row[4]  # Assuming the "HƯỚNG PHÒNG NGỪA" column is in the 5th position (column E)

        # Create a dictionary to store the extracted values
        maloi_data = {
            'maloi_maloi': maloi_maloi,
            'maloi_tenloi': maloi_tenloi,
            'maloi_khacphuc': maloi_khacphuc,
            'maloi_nguyennhan': maloi_nguyennhan,
            'maloi_phongngua': maloi_phongngua
        }

        # Append the dictionary to the data list
        data.append(maloi_data)

    # Iterate over the extracted data and save it to the database
    for maloi_data in data:
        maloi = Maloi(**maloi_data)
        db.add(maloi)

    db.commit()

    maloi_list = db.query(Maloi).order_by(Maloi.maloi_id.desc()).limit(50).all()
    return templates.TemplateResponse("maloi_xem.html", {"request": request, "maloi_list": maloi_list})


# MA LOI THEM
@app.get("/maloi/themui", response_class=HTMLResponse, tags=["MÃ LỖI"])
async def maloi_them_ui(request: Request):
    return templates.TemplateResponse("maloi_them.html", {"request": request})


@app.post("/maloi/them", response_class=HTMLResponse, tags=["MÃ LỖI"])
def maloi_them(request: Request, maloi_maloi: str = Form(...), maloi_tenloi: str = Form(...),
               maloi_khacphuc: str = Form(...),
               maloi_nguyennhan: str = Form(...), maloi_phongngua: str = Form(...), db: Session = Depends(get_db)):
    maloi = Maloi(maloi_maloi=maloi_maloi, maloi_tenloi=maloi_tenloi, maloi_khacphuc=maloi_khacphuc,
                  maloi_nguyennhan=maloi_nguyennhan, maloi_phongngua=maloi_phongngua)
    db.add(maloi)
    db.commit()
    time.sleep(1)
    maloi_list = db.query(Maloi).order_by(Maloi.maloi_id.desc()).limit(50).all()
    return templates.TemplateResponse("maloi_xem.html", {"request": request, "maloi_list": maloi_list})


# BAOCAOSUCO THEM
@app.get("/baocaosuco/themui", response_class=HTMLResponse, tags=["BÁO CÁO SỰ CỐ"])
async def baocaosuco_them_ui(request: Request, db: Session = Depends(get_db)):
    donhang_list = db.query(Donhang).all()
    maloi_list = db.query(Maloi).all()
    return templates.TemplateResponse("baocaosuco_them.html", {"request": request,
                                                               "donhang_list": donhang_list, "maloi_list": maloi_list})


# BAOCAOSUCO XEM
@app.post("/baocaosuco/them", tags=["BÁO CÁO SỰ CỐ"])
async def baocaosuco_them(request: Request):
    # Get the form data
    form_data = await request.form()

    # Extract the form input values
    madh = form_data["madh"]
    masp = form_data["masp"]
    maloi = form_data["maloi"]
    tenloi = form_data["tenloi"]
    soluongloi = form_data["soluongloi"]
    khacphuc = form_data["khacphuc"]
    nguyennhan = form_data["nguyennhan"]
    phongngua = form_data["phongngua"]
    ccemail = form_data["ccemail"]

    # Generate the PDF content
    pdf_content = f"""
    <h1>Báo Cáo Sự Cố</h1>
    <p>Mã Đơn Hàng: {madh}</p>
    <p>Mã Sản Phẩm: {masp}</p>
    <p>Mã Lỗi: {maloi}</p>
    <p>Tên Lỗi: {tenloi}</p>
    <p>Số Lượng Lỗi: {soluongloi}</p>
    <p>Hướng Khắc Phục: {khacphuc}</p>
    <p>Nguyên Nhân: {nguyennhan}</p>
    <p>Hướng Phòng Ngừa: {phongngua}</p>
    <p>CC Email: {ccemail}</p>
    """

    # Generate the PDF file name
    today = date.today().strftime("%Y_%m_%d")
    file_name = f"{madh}_{masp}_{maloi}.pdf"
    file_path = f"/storage/baocaosuco/{today}/{file_name}"

    # Generate the PDF and save it
    pdfkit.from_string(pdf_content, file_path, options={"quiet": ""})

    return {"message": "PDF generated and saved successfully"}


# FAULT READ
@app.get("/fault/report/view/admin", response_class=HTMLResponse, tags=["FAULT REPORT"])
def read_all_fault(request: Request, db: Session = Depends(get_db)):
    result = db.query(Fault).all()
    return templates.TemplateResponse("fault_report_view_list_admin.html", {"request": request, "fault_list": result})


@app.get("/fault/report/view", response_class=HTMLResponse, tags=["FAULT REPORT"])
def read_all_fault(request: Request, db: Session = Depends(get_db)):
    result = db.query(Fault).all()
    return templates.TemplateResponse("fault_report_view_list.html", {"request": request, "fault_list": result})


@app.get("/fault/report/view/{id}", response_class=HTMLResponse, tags=["FAULT REPORT"])
def read_fault(request: Request, id: int, db: Session = Depends(get_db)):
    result = db.query(Fault).filter(Fault.fault_id == id).first()
    return templates.TemplateResponse("fault_report_view_id.html", {"request": request, "fault": result})


# FAULT READ
# @app.get("/fault/view/admin", response_class=HTMLResponse)
# def read_all_fault_admin(request: Request, current_user=Depends(get_current_user), db: Session = Depends(get_db)):
#     if current_user.user_role != "admin":
#         raise HTTPException(status_code=403, detail="Forbidden")
#     result = db.query(Fault).all()
#     return templates.TemplateResponse("fault_report_view_list_admin.html", {"request": request, "fault_list": result})

# FAULT CREATE
@app.get("/fault/report/createui", response_class=HTMLResponse, tags=["FAULT REPORT"])
async def create_fault_ui(request: Request):
    return templates.TemplateResponse("fault_report_new.html", {"request": request})


@app.post("/fault/report/create", response_class=HTMLResponse, tags=["FAULT REPORT"])
def create_fault(request: Request, faultId: str = Form(...), faultName: str = Form(...), faultImage: str = Form(...),
                 faultDescription: str = Form(...), db: Session = Depends(get_db)):
    fault = Fault(fault_id=faultId, fault_name=faultName, fault_image=faultImage, fault_description=faultDescription)
    db.add(fault)
    db.commit()
    time.sleep(1)
    result = db.query(Fault).all()
    return templates.TemplateResponse("fault_report_view_list.html", {"request": request, "fault_list": result})


# FAULT DELETE
@app.get("/fault/report/delete/{id}", response_class=HTMLResponse, tags=["FAULT REPORT"])
def delete_fault(id: int, response: Response, request: Request, db: Session = Depends(get_db)):
    db.query(Fault).filter(Fault.fault_id == id).delete()
    db.commit()
    time.sleep(1)
    result = db.query(Fault).all()
    return templates.TemplateResponse("fault_report_view_list.html", {"request": request, "fault_list": result})


# FAULT UPDATE
@app.get("/fault/report/edit/{id}", response_class=HTMLResponse, tags=["FAULT REPORT"])
def edit_fault(request: Request, id: int, db: Session = Depends(get_db)):
    result = db.query(Fault).filter(Fault.fault_id == id).first()
    return templates.TemplateResponse("fault_report_edit.html", {"request": request, "fault": result})


@app.post("/fault/report/edit/{id}", response_class=HTMLResponse, tags=["FAULT REPORT"])
def update_fault(request: Request, id: int, faultName: str = Form(...), faultImage: str = Form(...),
                 faultDescription: str = Form(...), faultApproval: str = Form(...),
                 db: Session = Depends(get_db)):
    db.query(Fault).filter(Fault.fault_id == id).update({
        Fault.fault_name: faultName,
        Fault.fault_image: faultImage,
        Fault.fault_description: faultDescription,
        Fault.fault_approval: faultApproval
    })
    db.commit()
    time.sleep(1)
    result = db.query(Fault).all()
    return templates.TemplateResponse("fault_report_view_list.html", {"request": request, "fault_list": result})


# FAULT SOLUTION READ
@app.get("/fault/solution/view", response_class=HTMLResponse, tags=["FAULT SOLUTION"])
def read_all_fault_solution(request: Request, db: Session = Depends(get_db)):
    result = db.query(Solution).all()
    return templates.TemplateResponse("fault_solution_view_list.html",
                                      {"request": request, "faultsolution_list": result})


# FAULT SOLUTION CREATE
@app.get("/fault/solution/createui", response_class=HTMLResponse, tags=["FAULT SOLUTION"])
async def create_solution_ui(request: Request):
    return templates.TemplateResponse("fault_solution_new.html", {"request": request})


@app.post("/fault/solution/create", response_class=HTMLResponse, tags=["FAULT SOLUTION"])
def create_solution(request: Request, faultId: str = Form(...), solutionId: str = Form(...),
                    solutionDetail: str = Form(...), db: Session = Depends(get_db)):
    solution = Solution(fault_id=faultId, solution_id=solutionId, solution_detail=solutionDetail)
    db.add(solution)
    db.commit()
    time.sleep(1)
    result = db.query(solution).all()
    return templates.TemplateResponse("fault_solution_view_list.html", {"request": request, "solution_list": result})


# FAULT SOLUTION DELETE
@app.get("/fault/solution/delete/{id}", response_class=HTMLResponse, tags=["FAULT SOLUTION"])
def delete_solution(id: int, response: Response, request: Request, db: Session = Depends(get_db)):
    db.query(Solution).filter(Solution.solution_id == id).delete()
    db.commit()
    time.sleep(1)
    result = db.query(Solution).all()
    return templates.TemplateResponse("fault_solution_view_list.html", {"request": request, "solution_list": result})


# FAULT SOLUTION UPDATE
@app.get("/fault/solution/edit/{id}", response_class=HTMLResponse, tags=["FAULT SOLUTION"])
def edit_solution(request: Request, id: int, db: Session = Depends(get_db)):
    result = db.query(Solution).filter(Solution.solution_id == id).first()
    return templates.TemplateResponse("fault_solution_edit.html", {"request": request, "solution": result})


@app.post("/fault/solution/edit/{id}", response_class=HTMLResponse, tags=["FAULT SOLUTION"])
def update_solution(request: Request, id: int, faultId: str = Form(...), solutionDetail: str = Form(...),
                    db: Session = Depends(get_db)):
    db.query(Solution).filter(Solution.solution_id == id).update({
        Solution.fault_id: faultId,
        Solution.solution_detail: solutionDetail,
    })
    db.commit()
    time.sleep(1)
    result = db.query(Solution).all()
    return templates.TemplateResponse("fault_solution_view_list.html", {"request": request, "solution_list": result})


# STATISTICAL VIEW
@app.get("/statistical/view", response_class=HTMLResponse, tags=["STATISTICAL"])
def read_all_statistical(request: Request):
    return templates.TemplateResponse("statistical_view.html", {"request": request, })


@app.get("/statistical/fault/report/form", response_class=HTMLResponse, tags=["STATISTICAL"])
def read_fault_report(request: Request, db: Session = Depends(get_db)):
    result = db.query(Fault).all()
    return templates.TemplateResponse("statistical_fault_report.html", {"request": request, "fault_list": result})


@app.get("/statistical/fault/report/print", response_class=StreamingResponse, tags=["STATISTICAL"])
def generate_fault_report_pdf(request: Request, db: Session = Depends(get_db)):
    fault_data = db.query(Fault).all()
    solution_data = db.query(Solution).all()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)

    # Define table data and styles
    table_data = [
        ["Id", "Name", "Image", "Description", "Solution", "Approval"],
    ]
    for fault, solution in zip(fault_data, solution_data):
        table_data.append([
            fault.fault_id,
            fault.fault_name,
            fault.fault_image,
            fault.fault_description,
            solution.solution_detail,
            fault.fault_approval,
        ])

    table_style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
        ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
    ])

    table = Table(table_data)
    table.setStyle(table_style)

    # Build the PDF structure
    elements = []
    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    return StreamingResponse(BufferedReader(buffer), media_type="application/pdf",
                             headers={"Content-Disposition": "attachment; filename=fault_report.pdf"})
